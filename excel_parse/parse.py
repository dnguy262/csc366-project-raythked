import openpyxl


def parseProfileCharacteristics():
    path = "./ProfileCharacteristics.xlsx"

    wb_obj = openpyxl.load_workbook(
        path
    )  # To open the workbook, workbook object is created
    sheet_obj = (
        wb_obj.active
    )  # Get workbook active sheet object from the active attribute

    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    queryList = []
    for row in range(2, max_row + 1):  # skip column titles

        rowValues = []
        for col in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=row, column=col)
            rowValues.append(str(cell_obj.value))

        query = (
            """INSERT INTO JobDescriptors (Id, Dimension, Characteristic, Description) """
            """VALUES ({}, "{}", "{}", "{}");""".format(
                rowValues[0], rowValues[1], rowValues[2], rowValues[3]
            )
        )

        queryList.append(query)

        rowValues.clear()
    return queryList


def parseWorksheet():
    # Definition Worksheet
    path = "./Data-v03.xlsx"
    wb_obj = openpyxl.load_workbook(
        path
    )  # To open the workbook, workbook object is created

    # Aggregating insert queries by worksheet
    queries = {}
    for worksheet in wb_obj.sheetnames:
        sheet_obj = wb_obj[
            worksheet
        ]  # Get workbook active sheet object from the active attribute

        if worksheet == "URE Experience":
            parseUREExperienceSheet(sheet_obj, queries)

        else:
            max_col = sheet_obj.max_column
            max_row = sheet_obj.max_row

            for row in range(2, max_row + 1):  # skip column titles

                rowValues = []
                for col in range(1, max_col + 1):
                    cell_obj = sheet_obj.cell(row=row, column=col)
                    rowValues.append(str(cell_obj.value))

                # Some worksheets start in 2nd or 3rd row, this will catch empty rows
                if rowValues[0] == "None":
                    continue

                query = ""
                if worksheet == "Surveys":
                    query = f"""INSERT INTO Surveys (Title, Name, Description) VALUES ("{rowValues[1]}", "{rowValues[2]}", "{rowValues[3]}");"""

                elif worksheet == "Users":
                    query = f"""INSERT INTO UserAccounts (UserName, Notes, Category) VALUES ("{rowValues[1]}", "{rowValues[2]}", "Student");"""

                elif worksheet == "Survey Questions New":
                    query = f"""INSERT INTO Questions (SurveyId, QNumber, Text, Type, JobDescriptor) VALUES ({rowValues[0]}, {rowValues[1]}, "{rowValues[2]}", {rowValues[3]}, "{rowValues[4]}");"""

                elif worksheet == "QuestionResponses":
                    query = f"""INSERT INTO Choices (SurveyId, QNumber, CNumber, Value) VALUES ({rowValues[0]}, {rowValues[1]}, {rowValues[2]}, "{rowValues[3]}");"""

                # Replacing Python's None to MySQL's Null for missing fields
                query = query.replace('"None"', "Null")
                query = query.replace("None", "Null")

                if worksheet not in queries.keys():
                    queries[worksheet] = []

                # There are still other worksheets to be implemented
                if query:
                    queries[worksheet].append(query)

    # Hardcoding Missing Choices from Excel (12-94) have same choices
    cnumbers = range(8)
    cvalues = [
        "Prefer not to answer",
        "strongly disagree",
        "disagree",
        "somewhat disagree",
        "neither agree nor disagree",
        "somewhat agree",
        "agree",
        "strongly agree",
    ]
    questions_missing = range(13, 95)
    survey_id = 1 # URE Survey
    for qnumber in questions_missing:
        for cnumber, cvalue in zip(cnumbers, cvalues):
            query = f"""INSERT INTO Choices (SurveyId, QNumber, CNumber, Value) VALUES ({survey_id}, {qnumber}, {cnumber}, "{cvalue}");"""
            queries["QuestionResponses"].append(query)
    return queries


# RESOURCES
# https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/
# https://blog.piinalpin.com/2020/12/sql-generator/

"""
    This function updates the queries dictionary, and does not return anything
"""


def parseUREExperienceSheet(sheet_obj, queries):

    # initialize sheet queries list
    if "URE Experience" not in queries.keys():
        queries["URE Experience"] = []

    # hard coded max_col and max_row because rows and columns numbers were inaccurat, not sure why
    max_col = 96
    max_row = 27

    # None as of now... need to create surveySubmissions (?)
    surveySubmissionId = "Null"
    for row in range(3, max_row + 1):
        rowValues = []
        questionId = 1
        query = ""
        for col in range(3, max_col + 1):
            cell_obj = sheet_obj.cell(row=row, column=col)
            rowValues.append(str(cell_obj.value))
            query = f"""INSERT INTO Responses (SurveySubmissionId, QuestionId, SelectedChoiceId) VALUES ({surveySubmissionId}, {questionId}, {cell_obj.value});"""
            questionId += 1
            queries["URE Experience"].append(query)


# def parseUREExperienceSheet(sheet_obj, queries)

#     # initialize sheet queries list
#     if "URE Experience" not in queries.keys():
#         queries["URE Experience"] = []

#     # create mapping from question column number to questionIds
#     max_col = sheet_obj.max_column
#     max_row = sheet_obj.max_row
#     colNoToId = {}
#     for col in range(2, max_row + 1):
#         cell_obj = sheet_obj.cell(row = 1, column = col) # header row

#         # query sql with questionText to get questionId
#         questionText = str(cell_obj.value)
#         selectQuery = f"""
#                        SELECT QNumber
#                        FROM Questions
#                        WHERE Text = '{questionText}' and SurveyId = {1};
#                        """ # UREExperience --> SurveyId = 1

#         questionId = executeSelectQuery(selectQuery) # TODO: finish executeSelectQuery function
#         colNoToId[col] = questionId # mapping from col number to questionId

#     # insert survey submissions
#     for row in range(2, max_row + 1): # skip first two rows, each row is a survey submission

#         surveyId = 1 # since this is URE survey
#         query = f"""INSERT INTO SurveySubmissions (Id?, SurveyId)
#                     VALUES ({}, {surveyId});"""
#         # TODO: do i create survey submission id here? or does mysql auto generate it because of auto increment
#         # TODO: if SurveySubmissionId is auto-incremented, need to retrieve it in order to create Responses table
#         queries["URE Experience"].append[query]

#         # insert responses
#         for col in range(2, max_col + 1): # skip first two columns

#             surveySubmissionId = None # TODO
#             questionId = colNoToId[col]
#             selectedChoiceId = cell_obj.value # TODO: does this need to be cast as int?
#             query = f"""INSERT INTO Responses (SurveySubmissionId, QuestionId, SelectedChoiceId)
#                         VALUES ({surveySubmissionId}, {questionId}, {selectedChoiceId});"""
#             queries["URE Experience"].append[query]


# def executeSelectQuery(query):
#     connection = pymysql.connect(
#                 user     = "group5a",
#                 password = "_}2nSW6%?3hyr9Z",
#                 host     = "mysql.labthreesixfive.com",
#                 db       = "group5a",
#                 port     = 3306
#     )

#     with connection.cursor() as cursor:
#         cursor.execute(query)

#     # TODO: add way to return the results from a select query

#     connection.commit()
#     connection.close()


def main():
    definitions_queries = parseProfileCharacteristics()
    queries = parseWorksheet()

    all_queries = definitions_queries
    for queries in queries.values():
        all_queries.extend(queries)

    all_queries = [q + "\n" for q in all_queries]

    with open("insert_stms_excel.sql", "w") as file:
        file.writelines(all_queries)


if __name__ == "__main__":
    main()
