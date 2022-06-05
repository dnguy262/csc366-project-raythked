import openpyxl

def parseWorksheet():
    # Definition Worksheet
    path = './Data-v03.xlsx'
    wb_obj = openpyxl.load_workbook(path) # To open the workbook, workbook object is created
    
    # Aggregating insert queries by worksheet
    queries = {}
    for worksheet in wb_obj.sheetnames:
        sheet_obj = wb_obj[worksheet] # Get workbook active sheet object from the active attribute
        
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row

        for row in range(3, max_row + 1): # skip column titles
            
            rowValues = []
            for col in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row = row, column = col)
                rowValues.append(str(cell_obj.value))
            
            # Some worksheets start in 2nd or 3rd row, this will catch empty rows
            if rowValues[0] == 'None':
                continue
            
            query = ""

            if worksheet == "URE Experience":
                print(max_row)
                print(rowValues)
                numOfQuestions = 94
                for i in range(1,numOfQuestions+1): 
                    print(i)
                

            # Replacing Python's None to MySQL's Null for missing fields
            query = query.replace('"None"', "Null")
            query = query.replace('None', "Null")
            
            if worksheet not in queries.keys():
                queries[worksheet] = []
            
            # There are still other worksheets to be implemented
            if query:
                queries[worksheet].append(query)
    return queries

parseWorksheet()