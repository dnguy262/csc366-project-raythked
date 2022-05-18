import openpyxl
 
def parseProfileCharacteristics():
    path = './ProfileCharacteristics.xlsx'
 
    wb_obj = openpyxl.load_workbook(path) # To open the workbook, workbook object is created
    sheet_obj = wb_obj.active # Get workbook active sheet object from the active attribute
    
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    queryList = []
    for row in range(2, max_row + 1): # skip column titles
        
        rowValues = []
        for col in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row = row, column = col)
            rowValues.append(str(cell_obj.value))

        query = ("""INSERT INTO JobDescriptors (Id, Dimension, Characteristic, Description) """
                """VALUES ({}, "{}", "{}", "{}");""".format(rowValues[0], rowValues[1], rowValues[2], rowValues[3]))

        queryList.append(query)

        rowValues.clear()
    return queryList

def parseWorksheet():
    # Definition Worksheet
    path = './Data-v03.xlsx'
    wb_obj = openpyxl.load_workbook(path) # To open the workbook, workbook object is created
    
    # Aggregating insert queries by worksheet
    queries = {}
    for worksheet in wb_obj.sheetnames:
        sheet_obj = wb_obj[worksheet] # Get workbook active sheet object from the active attribute
        
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row

        for row in range(2, max_row + 1): # skip column titles
            
            rowValues = []
            for col in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row = row, column = col)
                rowValues.append(str(cell_obj.value))
            
            # Some worksheets start in 2nd or 3rd row, this will catch empty rows
            if not rowValues[0]:
                continue
            
            query = ""
            if worksheet == "Surveys":
                query = f"""INSERT INTO Surveys (Title, Name, Description) VALUES ("{rowValues[1]}", "{rowValues[2]}", "{rowValues[3]}");"""

            elif worksheet == "Users":
                query = f"""INSERT INTO UserAccounts (UserName, Notes, Category) VALUES ("{rowValues[1]}", "{rowValues[2]}", "Student");"""

            elif worksheet == "Survey Questions New":
                query = f"""INSERT INTO Questions (SurveyId, QNumber, Text, Type, JobDescriptor) VALUES ({rowValues[0]}, {rowValues[1]}, "{rowValues[2]}", {rowValues[3]}, "{rowValues[4]}");"""

            elif worksheet == "QuestionResponses":
                query = f"""INSERT INTO Choices (SurveyId, QuestionId, CNumber, Value) VALUES ({rowValues[0]}, {rowValues[1]}, {rowValues[2]}, "{rowValues[3]}");"""

            # Replacing Python's None to MySQL's Null for missing fields
            query = query.replace('"None"', "Null")
            
            if worksheet not in queries.keys():
                queries[worksheet] = []
            
            # There are still other worksheets to be implemented
            if query:
                queries[worksheet].append(query)

    return queries

# print(queryList)

# RESOURCES
# https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/
# https://blog.piinalpin.com/2020/12/sql-generator/

def main():
    definitions_queries = parseProfileCharacteristics()
    queries = parseWorksheet()
    
    # Comment this out for debugging
    # for k, v in queries.items():
    #     print(k,v)
if __name__ == '__main__':
    main()