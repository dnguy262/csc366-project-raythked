from bs4 import BeautifulSoup
import pymysql
import openpyxl
import requests
import base64
import math
from concurrent.futures import ThreadPoolExecutor, as_completed

from manager import executeQuery

# Executes any Select Query
def executeSelect(query):
    connection = pymysql.connect(
                user     = "group5a",
                password = "_}2nSW6%?3hyr9Z",
                host     = "mysql.labthreesixfive.com",
                db       = "group5a",
                port     = 3306
    )
    with connection.cursor() as cursor:
        cursor.execute(query)
    connection.commit()
    return cursor.fetchall()

# Gets ONET descriptor codes for each job descriptor
def get_characteristic_codes():
    path = './onet_mapping.xlsx'
 
    wb_obj = openpyxl.load_workbook(path) # To open the workbook, workbook object is created
    sheet_obj = wb_obj.active # Get workbook active sheet object from the active attribute
    
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    profile_characteristic_codes = {}
    for row in range(7, max_row + 1): # skip column titles
        
        descriptor_codes = []
        profile_characteristic = ''
        for col in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row = row, column = col)

            if col == 1:
                profile_characteristic = cell_obj.value

            if cell_obj.hyperlink:

                # Get ONET descriptor code from link
                descriptor_link = cell_obj.hyperlink.target
                descriptor_code = descriptor_link.split("/")[-1]
                if descriptor_code:
                    descriptor_codes.append(descriptor_code)
        
        if descriptor_codes:

            # Last character is the profile characteristic id
            profile_characteristic_codes[profile_characteristic.split()[-1]] = descriptor_codes
    return profile_characteristic_codes

# Selects all profile ids and job codes from ONET Profile
def get_onet_profiles():
    query = """SELECT Id, Code FROM Profiles WHERE Category = 'ONET' """
    ids, codes = zip(*executeSelect(query))
    return ids, codes

def fetch_onet_job(profile_id, job_code, profile_characteristic_codes):
    queries = []

    # Pinging ONET API 
    api_endpoint = f"https://services.onetcenter.org/ws/online/occupations/{job_code}/details/"
    username = 'raythked'
    password =  '5638nqd'
    headers = {
        'Authorization': 'Basic ' + base64.standard_b64encode((username + ':' + password).encode()).decode(),
    }

    response = requests.get(api_endpoint, headers=headers)
    soup = BeautifulSoup(response.content, "lxml")


    for descriptor_id, descriptor_codes in profile_characteristic_codes.items():
        scores = []
        for descriptor_code in descriptor_codes:
            element = soup.find(id=descriptor_code, recursive=True)
            if element:

                # Get Level Score, if present
                score = element.find('score', attrs={'scale': 'Level'})

                # Get Context Score, if present
                if not score:
                    score = element.find('score', attrs={'scale': 'Context'})

                # Get Importance Score, if present
                if not score:
                    score = element.find('score', attrs={'scale': 'Importance'})

                if not score:
                    continue

                scores.append(int(score.text))
        
        if scores:
            score = math.ceil(max(scores) * 7 / 100)
            query = f"""("{score}", NULL, "{descriptor_id}", "{profile_id}"),"""
            queries.append(query)
    
    return queries

def main():
    profile_characteristic_codes = get_characteristic_codes()
    profile_ids, job_codes = get_onet_profiles()
    queries = ["INSERT INTO DescriptorScores (Score, Importance, JobDescriptorId, ProfileId) VALUES "]

    # Fetches each onet job in a thread
    print("fetching onet job descriptors ...")
    with ThreadPoolExecutor() as executor:
        for job_queries in executor.map(fetch_onet_job, profile_ids, job_codes, [profile_characteristic_codes]*len(profile_ids)):
            queries.extend(job_queries)
    
    print("done")
    queries[-1] = queries[-1][:-1] + ";"
    queries = [q + "\n" for q in queries]
    
    # Creates sql file
    with open('insert_stms_onet.sql', 'w') as file:
        file.writelines(queries)

if __name__ == '__main__':
    main()